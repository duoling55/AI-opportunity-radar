from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from opportunity_radar.export.excel import HEADERS, ExportRow, export_workbook


def _row(sheet_name: str = "重点商机") -> ExportRow:
    return ExportRow(
        sheet_name,
        {
            "商机等级": "A",
            "商机评分": 88,
            "政策名称": "设备更新通知",
            "发布日期": date(2026, 7, 20),
            "国标行业大类代码": "C34",
            "国标行业大类名称": "通用设备制造业",
            "业务行业标签": "通用装备制造",
            "机会场景": "设备更新",
            "政策原文依据": "支持设备更新",
            "行业营销开场白": "交流设备更新安排",
            "政策原文链接": "https://example.gov.cn/policy",
            "附件链接": "https://example.gov.cn/attachment.pdf",
        },
    )


def test_export_has_two_required_sheets_and_preserves_existing_file(tmp_path: Path) -> None:
    first = export_workbook([_row()], tmp_path, date(2026, 7, 29))
    second = export_workbook([_row()], tmp_path, date(2026, 7, 29))

    assert first != second
    assert first.exists()
    assert second.exists()
    assert load_workbook(first).sheetnames == ["重点商机", "政策观察"]


def test_export_writes_all_business_headers_and_sheet_specific_rows(tmp_path: Path) -> None:
    path = export_workbook([_row(), _row("政策观察")], tmp_path, date(2026, 7, 29))
    workbook = load_workbook(path)

    assert len(HEADERS) == 36
    assert [cell.value for cell in workbook["重点商机"][1]] == HEADERS
    assert workbook["重点商机"].max_row == 2
    assert workbook["政策观察"].max_row == 2


def test_export_freezes_and_filters_headers_and_creates_clickable_links(tmp_path: Path) -> None:
    path = export_workbook([_row()], tmp_path, date(2026, 7, 29))
    sheet = load_workbook(path)["重点商机"]

    source = sheet.cell(2, HEADERS.index("政策原文链接") + 1)
    attachment = sheet.cell(2, HEADERS.index("附件链接") + 1)
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == f"A1:AJ{sheet.max_row}"
    assert source.hyperlink is not None
    assert source.hyperlink.target == "https://example.gov.cn/policy"
    assert attachment.hyperlink is not None
    assert attachment.hyperlink.target == "https://example.gov.cn/attachment.pdf"


def test_export_sorts_by_score_then_publish_date_and_writes_true_dates(
    tmp_path: Path,
) -> None:
    older_high = ExportRow(
        "重点商机",
        {**_row().values, "政策名称": "高分较早", "商机评分": 90, "发布日期": date(2026, 7, 20)},
    )
    newer_high = ExportRow(
        "重点商机",
        {**_row().values, "政策名称": "高分较新", "商机评分": 90, "发布日期": date(2026, 7, 21)},
    )
    lower = ExportRow(
        "重点商机",
        {**_row().values, "政策名称": "低分", "商机评分": 70, "发布日期": date(2026, 7, 29)},
    )

    path = export_workbook([lower, older_high, newer_high], tmp_path, date(2026, 7, 29))
    sheet = load_workbook(path)["重点商机"]
    title_column = HEADERS.index("政策名称") + 1
    date_column = HEADERS.index("发布日期") + 1

    assert [sheet.cell(row, title_column).value for row in range(2, 5)] == [
        "高分较新",
        "高分较早",
        "低分",
    ]
    assert sheet.cell(2, date_column).value.date() == date(2026, 7, 21)
    assert sheet.cell(2, date_column).number_format == "yyyy-mm-dd"


def test_export_applies_grade_fill_wrap_and_reasonable_column_widths(
    tmp_path: Path,
) -> None:
    rows = [
        ExportRow("重点商机", {**_row().values, "商机等级": "A"}),
        ExportRow("重点商机", {**_row().values, "商机等级": "B"}),
        ExportRow("政策观察", {**_row("政策观察").values, "商机等级": "观察"}),
    ]

    path = export_workbook(rows, tmp_path, date(2026, 7, 29))
    workbook = load_workbook(path)
    priority = workbook["重点商机"]
    observation = workbook["政策观察"]
    summary_column = HEADERS.index("政策摘要") + 1
    score_column = HEADERS.index("商机评分") + 1

    assert priority["A2"].fill.fgColor.rgb != priority["A3"].fill.fgColor.rgb
    assert observation["A2"].fill.fill_type == "solid"
    assert priority.cell(2, summary_column).alignment.wrap_text is True
    assert priority.column_dimensions["D"].width > priority.column_dimensions["B"].width
    assert priority.cell(2, score_column).alignment.wrap_text is not True


def test_export_keeps_multiple_attachment_urls_and_uses_valid_first_target(
    tmp_path: Path,
) -> None:
    row = ExportRow(
        "重点商机",
        {
            **_row().values,
            "附件链接": (
                "https://example.gov.cn/one.pdf、https://example.gov.cn/two.docx"
            ),
        },
    )

    path = export_workbook([row], tmp_path, date(2026, 7, 29))
    cell = load_workbook(path)["重点商机"].cell(2, HEADERS.index("附件链接") + 1)

    assert cell.value == (
        "https://example.gov.cn/one.pdf\nhttps://example.gov.cn/two.docx"
    )
    assert cell.hyperlink is not None
    assert cell.hyperlink.target == "https://example.gov.cn/one.pdf"
