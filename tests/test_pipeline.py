from __future__ import annotations

import json
from dataclasses import replace
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from opportunity_radar.compliance import (
    ComplianceAuditSnapshot,
    ComplianceSource,
    RateLimitPolicy,
)
from opportunity_radar.config import RunConfig, SourceConfig
from opportunity_radar.http import OfficialHttpClient
from opportunity_radar.models import (
    Evidence,
    IndustryOpportunity,
    PolicyAnalysis,
    PolicyCandidate,
    PolicyDocument,
)
from opportunity_radar.parsing.html import DocumentRetriever
from opportunity_radar.quality.scripts import DISCLAIMER
from opportunity_radar.sources.base import GenericHtmlSource
from opportunity_radar.state import StateStore


def _candidate(title: str) -> PolicyCandidate:
    return PolicyCandidate(
        source_id="good",
        title=title,
        detail_url=f"https://example.gov.cn/{title}",
        published_at=date(2026, 7, 10),
    )


def _document(title: str, **changes: object) -> PolicyDocument:
    values: dict[str, object] = {
        "policy_id": title,
        "source_id": "good",
        "source_name": "测试政策源",
        "region": "全国",
        "title": title,
        "detail_url": f"https://example.gov.cn/{title}",
        "publisher": "测试发布机构",
        "publish_date": date(2026, 7, 10),
        "raw_text": "支持设备更新和技术改造。",
        "normalized_text": "支持设备更新和技术改造。",
        "collected_at": datetime(2026, 7, 29, tzinfo=UTC),
        "content_hash": f"hash-{title}",
        "snapshot_path": f"raw/{title}.html",
    }
    values.update(changes)
    return PolicyDocument(**values)


def _opportunity(**changes: object) -> IndustryOpportunity:
    values: dict[str, object] = {
        "section_code": "C",
        "section_name": "制造业",
        "division_code": "C34",
        "division_name": "通用设备制造业",
        "business_tags": ["通用装备制造"],
        "confidence": 0.9,
        "scenarios": ["设备更新"],
        "evidence": [Evidence(quote="支持设备更新", location="正文字符1-6")],
        "leasing_relevance": "设备可通过融资租赁配置",
        "recommended_action": "联系园区客户经理",
        "opening_script": "我们可以交流近期设备更新安排",
    }
    values.update(changes)
    return IndustryOpportunity(**values)


def _verified_compliance_source(source_id: str) -> ComplianceSource:
    return ComplianceSource(
        source_id=source_id,
        display_name=source_id.upper(),
        phase="verified",
        enabled=True,
        terms="书面许可允许按核验字段和限频自动访问。",
        terms_confirmed=True,
        registration="completed",
        registration_completed=True,
        authorization="written_permission",
        rate_limit=RateLimitPolicy(12, 60, 5),
        selected_data_scope="惠企政策数据集 v1",
        field_permission_confirmed=True,
        evidence_url=f"https://{source_id}.example.gov.cn/permission",
        verified_at=date(2026, 7, 1),
        review_due_at=date(2026, 9, 29),
        owner="policy-ops",
        available_fields=("标题",),
    )


class _Source:
    def __init__(self, candidates: list[PolicyCandidate] | None = None, *, fails: bool = False):
        self.candidates = candidates or []
        self.fails = fails

    def discover(self, start: date, end: date) -> list[PolicyCandidate]:
        if self.fails:
            raise RuntimeError("listing failed")
        return self.candidates


class _AttachmentSource(GenericHtmlSource):
    def __init__(self, config: SourceConfig, candidates: list[PolicyCandidate]) -> None:
        super().__init__(config, OfficialHttpClient(config))
        self.candidates = candidates

    def discover(self, start: date, end: date) -> list[PolicyCandidate]:
        return self.candidates


class _Retriever:
    def __init__(
        self,
        documents: dict[str, PolicyDocument],
        *,
        permission_failure: str | None = None,
    ):
        self.documents = documents
        self.permission_failure = permission_failure
        self.seen: list[str] = []

    def fetch_document(
        self,
        source: _Source,
        candidate: PolicyCandidate,
        collected_at: datetime,
        raw_dir: Path,
    ) -> PolicyDocument:
        self.seen.append(candidate.title)
        if candidate.title == self.permission_failure:
            raise PermissionError("source access restricted: 403")
        if candidate.title == "parse-failure":
            raise ValueError("detail page changed")
        return self.documents[candidate.title]


class _Analyzer:
    def __init__(self) -> None:
        self.seen: list[str] = []

    def analyze(
        self,
        document: PolicyDocument,
        valid_codes: set[str],
        business_tags: list[str],
    ) -> PolicyAnalysis:
        self.seen.append(document.title)
        if document.title == "analysis-failure":
            raise RuntimeError("model unavailable")
        if document.title == "benefit-without-opportunity":
            return PolicyAnalysis(
                is_benefit_policy=True,
                summary="惠企政策摘要",
                support_direction="支持设备更新",
            )
        if document.title == "quality-gated":
            return PolicyAnalysis(
                is_benefit_policy=True,
                summary="存在不合规营销表达",
                support_direction="支持设备更新",
                opportunities=[_opportunity(opening_script="政府背书，保证获批")],
            )
        return PolicyAnalysis(
            is_benefit_policy=True,
            summary="设备更新政策摘要",
            support_direction="支持设备更新",
            opportunities=[_opportunity()],
        )


def test_pipeline_isolates_failures_filters_and_exports_gated_rows(tmp_path: Path) -> None:
    from opportunity_radar.pipeline import run_pipeline

    titles = [
        "outside-window",
        "unchanged",
        "parse-failure",
        "analysis-failure",
        "benefit-without-opportunity",
        "quality-gated",
        "priority",
    ]
    candidates = [_candidate(title) for title in titles]
    documents = {title: _document(title) for title in titles if title != "parse-failure"}
    documents["outside-window"] = _document("outside-window", publish_date=date(2026, 6, 30))
    state_path = tmp_path / "state.sqlite3"
    seeded_store = StateStore(state_path)
    seeded_store.record_success("unchanged", "hash-unchanged")
    seeded_store.connection.close()
    analyzer = _Analyzer()
    config = RunConfig(
        date(2026, 7, 1),
        date(2026, 7, 30),
        ("good", "bad"),
        tmp_path,
        state_path,
        tmp_path / "raw",
        tmp_path / "normalized",
        (
            ComplianceAuditSnapshot(
                source_id="good",
                verified_at=date(2026, 7, 1),
                evidence_url="https://example.gov.cn/permission",
                adapter_version="1.2.3",
            ),
            ComplianceAuditSnapshot(
                source_id="bad",
                verified_at=date(2026, 7, 2),
                evidence_url="https://example.gov.cn/permission/bad",
                adapter_version="2.0.0",
            ),
        ),
    )

    workbook_path, report_path = run_pipeline(
        config,
        {"good": _Source(candidates), "bad": _Source(fails=True)},
        _Retriever(documents),
        analyzer,
        {"C", "C34"},
        ["通用装备制造"],
    )

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report == {
        "discovered": 7,
        "changed": 4,
        "skipped": 1,
        "source_failures": 1,
        "parse_failures": 1,
        "analysis_failures": 1,
        "priority_rows": 1,
        "observation_rows": 3,
        "compliance_audit": [
            {
                "source_id": "good",
                "verified_at": "2026-07-01",
                "evidence_url": "https://example.gov.cn/permission",
                "adapter_version": "1.2.3",
            },
            {
                "source_id": "bad",
                "verified_at": "2026-07-02",
                "evidence_url": "https://example.gov.cn/permission/bad",
                "adapter_version": "2.0.0",
            },
        ],
    }


    assert "outside-window" not in analyzer.seen
    assert "unchanged" not in analyzer.seen

    workbook = load_workbook(workbook_path)
    assert workbook.sheetnames == ["重点商机", "政策观察"]
    priority = workbook["重点商机"]
    observation = workbook["政策观察"]
    assert priority.max_row == 2
    assert observation.max_row == 4

    headers = {cell.value: cell.column for cell in priority[1]}
    all_rows = list(priority.iter_rows(min_row=2)) + list(observation.iter_rows(min_row=2))
    assert all(row[headers["免责声明"] - 1].value == DISCLAIMER for row in all_rows)
    assert all(
        DISCLAIMER in row[headers["行业营销开场白"] - 1].value
        for row in all_rows
        if row[headers["行业营销开场白"] - 1].value
    )

    observation_rows = {
        row[headers["政策名称"] - 1].value: row for row in observation.iter_rows(min_row=2)
    }
    assert (
        observation_rows["benefit-without-opportunity"][headers["复核原因"] - 1].value
        == "惠企政策未识别到可验证的机会行业"
    )
    assert (
        observation_rows["quality-gated"][headers["复核原因"] - 1].value
        == "营销话术含过度承诺或政府背书表达"
    )
    assert "AI 分析失败" in (observation_rows["analysis-failure"][headers["复核原因"] - 1].value)
    assert "政府背书" not in observation_rows["quality-gated"][headers["行业营销开场白"] - 1].value


def test_pipeline_force_reanalyze_bypasses_incremental_state(tmp_path: Path) -> None:
    from opportunity_radar.pipeline import run_pipeline

    document = _document("unchanged")
    state_path = tmp_path / "state.sqlite3"
    store = StateStore(state_path)
    store.record_success(document.policy_id, document.content_hash)
    store.connection.close()
    analyzer = _Analyzer()
    config = RunConfig(
        date(2026, 7, 1),
        date(2026, 7, 30),
        ("good",),
        output_dir=tmp_path / "outputs",
        state_path=state_path,
        raw_dir=tmp_path / "raw",
        normalized_dir=tmp_path / "normalized",
        force_reanalyze=True,
    )

    _, report_path = run_pipeline(
        config,
        {"good": _Source([_candidate("unchanged")])},
        _Retriever({"unchanged": document}),
        analyzer,
        {"C", "C34"},
        ["通用装备制造"],
    )

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert analyzer.seen == ["unchanged"]
    assert report["changed"] == 1
    assert report["skipped"] == 0


def test_pipeline_stops_source_after_detail_permission_failure(tmp_path: Path) -> None:
    from opportunity_radar.pipeline import run_pipeline

    candidates = [_candidate("restricted"), _candidate("must-not-fetch")]
    retriever = _Retriever(
        {
            "restricted": _document("restricted"),
            "must-not-fetch": _document("must-not-fetch"),
        },
        permission_failure="restricted",
    )
    config = RunConfig(
        date(2026, 7, 1),
        date(2026, 7, 30),
        ("good",),
        tmp_path,
        tmp_path / "state.sqlite3",
        tmp_path / "raw",
        tmp_path / "normalized",
    )

    _, report_path = run_pipeline(
        config,
        {"good": _Source(candidates)},
        retriever,
        _Analyzer(),
        {"C", "C34"},
        ["通用装备制造"],
    )

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert retriever.seen == ["restricted"]
    assert report["source_failures"] == 1
    assert report["parse_failures"] == 0


def test_pipeline_stops_source_after_attachment_permission_failure(
    httpx_mock, tmp_path: Path
) -> None:
    from opportunity_radar.pipeline import run_pipeline

    source_config = SourceConfig(
        "good",
        "测试政策源",
        "全国",
        (),
        ("example.gov.cn",),
        request_interval_seconds=0,
    )
    candidate = PolicyCandidate(
        source_id="good",
        title="受限附件政策",
        detail_url="https://example.gov.cn/policies/restricted.html",
        published_at=date(2026, 7, 10),
    )
    httpx_mock.add_response(
        url=str(candidate.detail_url),
        text=(
            "<article>支持设备更新。</article>"
            '<a href="/files/restricted.pdf">受限附件</a>'
            '<a href="/files/must-not-request.pdf">不得请求附件</a>'
        ),
    )
    httpx_mock.add_response(
        url="https://example.gov.cn/files/restricted.pdf", status_code=403
    )
    config = RunConfig(
        date(2026, 7, 1),
        date(2026, 7, 30),
        ("good",),
        tmp_path,
        tmp_path / "state.sqlite3",
        tmp_path / "raw",
        tmp_path / "normalized",
    )

    _, report_path = run_pipeline(
        config,
        {"good": _AttachmentSource(source_config, [candidate])},
        DocumentRetriever(),
        _Analyzer(),
        {"C", "C34"},
        ["通用装备制造"],
    )

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert [str(request.url) for request in httpx_mock.get_requests()] == [
        str(candidate.detail_url),
        "https://example.gov.cn/files/restricted.pdf",
    ]
    assert report["source_failures"] == 1
    assert report["parse_failures"] == 0


def test_pipeline_deduplicates_cross_source_repost_and_prefers_issuer(
    tmp_path: Path,
) -> None:
    from opportunity_radar.pipeline import run_pipeline

    government = _document(
        "government-repost",
        policy_id="government",
        source_id="government",
        source_name="江苏省人民政府政策库",
        publisher="江苏省工业和信息化厅",
        document_number="苏工信装〔2026〕1号",
        detail_url="https://example.gov.cn/government",
        content_hash="shared-hash",
    )
    issuer = _document(
        "issuer-original",
        policy_id="issuer",
        source_id="issuer",
        source_name="江苏省工业和信息化厅",
        publisher="江苏省工业和信息化厅",
        document_number="苏工信装[2026]1号",
        detail_url="https://example.gov.cn/issuer",
        content_hash="shared-hash",
    )
    config = RunConfig(
        date(2026, 7, 1),
        date(2026, 7, 30),
        ("government", "issuer"),
        tmp_path,
        tmp_path / "state.sqlite3",
        tmp_path / "raw",
        tmp_path / "normalized",
    )

    workbook_path, report_path = run_pipeline(
        config,
        {
            "government": _Source([_candidate("government-repost")]),
            "issuer": _Source([_candidate("issuer-original")]),
        },
        _Retriever(
            {
                "government-repost": government,
                "issuer-original": issuer,
            }
        ),
        _Analyzer(),
        {"C": "制造业", "C34": "通用设备制造业"},
        ["通用装备制造"],
    )

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["changed"] == 1
    assert report["skipped"] == 1
    priority = load_workbook(workbook_path)["重点商机"]
    headers = {cell.value: cell.column for cell in priority[1]}
    assert priority.cell(2, headers["数据来源"]).value == "江苏省工业和信息化厅"
    source_links = priority.cell(2, headers["政策原文链接"]).value
    assert source_links == (
        "https://example.gov.cn/issuer\nhttps://example.gov.cn/government"
    )


def test_fixture_run_writes_priority_and_observation_rows(
    tmp_path: Path,
    fixture_sources: dict[str, Any],
    fixture_retriever: Any,
    fixture_analyzer: Any,
) -> None:
    from opportunity_radar.pipeline import run_pipeline

    config = RunConfig(
        date(2026, 7, 1),
        date(2026, 7, 30),
        ("fixture",),
        tmp_path,
        tmp_path / "state.sqlite3",
        tmp_path / "raw",
        tmp_path / "normalized",
    )

    workbook_path, report_path = run_pipeline(
        config,
        fixture_sources,
        fixture_retriever,
        fixture_analyzer,
        {"C", "C34"},
        ["通用装备制造"],
    )

    workbook = load_workbook(workbook_path)
    assert workbook["重点商机"].max_row == 2
    assert workbook["政策观察"].max_row == 2
    assert '"analysis_failures": 0' in report_path.read_text(encoding="utf-8")


class _FailureAnalyzer:
    def __init__(self, failure_kind: str) -> None:
        self.failure_kind = failure_kind

    def analyze(
        self,
        document: PolicyDocument,
        valid_codes: set[str],
        business_tags: list[str],
    ) -> PolicyAnalysis:
        if self.failure_kind == "timeout":
            raise TimeoutError("analysis timed out")
        if self.failure_kind == "invalid-json":
            raise json.JSONDecodeError("invalid analysis JSON", "not-json", 0)
        PolicyAnalysis.model_validate({"is_benefit_policy": True})
        raise AssertionError("schema validation must fail")


@pytest.mark.parametrize(
    ("failure_kind", "error_type"),
    [
        ("timeout", "TimeoutError"),
        ("invalid-json", "JSONDecodeError"),
        ("schema", "ValidationError"),
    ],
)
def test_pipeline_exports_traceable_observation_when_analysis_fails(
    failure_kind: str,
    error_type: str,
    tmp_path: Path,
) -> None:
    from opportunity_radar.pipeline import run_pipeline

    document = _document(f"analysis-{failure_kind}")
    config = RunConfig(
        date(2026, 7, 1),
        date(2026, 7, 30),
        ("good",),
        tmp_path,
        tmp_path / "state.sqlite3",
        tmp_path / "raw",
        tmp_path / "normalized",
    )

    workbook_path, report_path = run_pipeline(
        config,
        {"good": _Source([_candidate(document.title)])},
        _Retriever({document.title: document}),
        _FailureAnalyzer(failure_kind),
        {"C", "C34"},
        ["通用装备制造"],
    )

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["analysis_failures"] == 1
    assert report["observation_rows"] == 1

    observation = load_workbook(workbook_path)["政策观察"]
    assert observation.max_row == 2
    values = {cell.value: observation.cell(2, cell.column).value for cell in observation[1]}
    assert values["政策名称"] == document.title
    assert values["政策原文链接"] == str(document.detail_url)
    assert values["数据来源"] == document.source_name
    assert values["发布机构"] == document.publisher
    assert document.normalized_text in values["政策摘要"]
    assert values["国标行业门类名称"] == "待复核"
    assert values["国标行业门类代码"] == "待复核"
    assert values["国标行业大类名称"] == "待复核"
    assert values["国标行业大类代码"] == "待复核"
    assert "AI 分析失败" in values["复核原因"]
    assert error_type in values["复核原因"]
    assert values["免责声明"] == DISCLAIMER


def test_cli_uses_manual_dates_sources_and_environment_ai_credentials(
    monkeypatch: Any, tmp_path: Path
) -> None:
    from opportunity_radar import cli

    configured = {
        source_id: SourceConfig(
            source_id,
            source_id.upper(),
            "全国",
            (f"https://{source_id}.example.gov.cn/list",),
            (f"{source_id}.example.gov.cn",),
            request_interval_seconds=5,
            adapter_version="1.0.0",
        )
        for source_id in ("miit", "ndrc", "unused")
    }
    captured: dict[str, object] = {}

    class _ConfiguredAnalyzer:
        def __init__(self, base_url: str, api_key: str, model: str) -> None:
            captured["credentials"] = (base_url, api_key, model)

    def fake_build_sources(selected: dict[str, SourceConfig]) -> dict[str, object]:
        captured["selected"] = tuple(selected)
        return {source_id: object() for source_id in selected}

    def fake_run_pipeline(config: RunConfig, *args: object) -> tuple[Path, Path]:
        captured["config"] = config
        return tmp_path / "result.xlsx", tmp_path / "report.json"

    monkeypatch.setenv("OPPORTUNITY_RADAR_LLM_BASE_URL", "https://llm.example/v1")
    monkeypatch.setenv("OPPORTUNITY_RADAR_LLM_API_KEY", "runtime-secret")
    monkeypatch.setenv("OPPORTUNITY_RADAR_LLM_MODEL", "runtime-model")
    monkeypatch.setattr(cli, "load_sources", lambda path: configured)
    monkeypatch.setattr(
        cli,
        "load_compliance_sources",
        lambda path: {source_id: _verified_compliance_source(source_id) for source_id in configured},
    )
    monkeypatch.setattr(cli, "build_sources", fake_build_sources)
    monkeypatch.setattr(cli, "load_industry_codes", lambda path: {"C", "C34"})
    monkeypatch.setattr(cli, "OpenAICompatibleAnalyzer", _ConfiguredAnalyzer)
    monkeypatch.setattr(cli, "run_pipeline", fake_run_pipeline)

    cli.main(
        [
            "run",
            "--start-date",
            "2026-07-01",
            "--end-date",
            "2026-07-30",
            "--sources",
            "miit,ndrc",
        ]
    )

    config = captured["config"]
    assert isinstance(config, RunConfig)
    assert (config.start_date, config.end_date) == (date(2026, 7, 1), date(2026, 7, 30))
    assert config.source_ids == ("miit", "ndrc")
    assert config.compliance_audit == (
        ComplianceAuditSnapshot(
            source_id="miit",
            verified_at=date(2026, 7, 1),
            evidence_url="https://miit.example.gov.cn/permission",
            adapter_version="1.0.0",
        ),
        ComplianceAuditSnapshot(
            source_id="ndrc",
            verified_at=date(2026, 7, 1),
            evidence_url="https://ndrc.example.gov.cn/permission",
            adapter_version="1.0.0",
        ),
    )
    assert captured["selected"] == ("miit", "ndrc")
    assert captured["credentials"] == (
        "https://llm.example/v1",
        "runtime-secret",
        "runtime-model",
    )


@pytest.mark.parametrize(
    ("base_url", "model", "expected_base_url", "expected_model"),
    [
        (None, None, "https://api.minimaxi.com/anthropic", "MiniMax-M3"),
        (
            "https://minimax-proxy.example/anthropic",
            "custom-minimax-model",
            "https://minimax-proxy.example/anthropic",
            "custom-minimax-model",
        ),
    ],
)
def test_cli_selects_minimax_anthropic_provider_with_defaults_and_overrides(
    monkeypatch: Any,
    tmp_path: Path,
    base_url: str | None,
    model: str | None,
    expected_base_url: str,
    expected_model: str,
) -> None:
    from opportunity_radar import cli

    configured = {
        "miit": SourceConfig(
            "miit",
            "MIIT",
            "全国",
            ("https://miit.example.gov.cn/list",),
            ("miit.example.gov.cn",),
            request_interval_seconds=5,
            adapter_version="1.0.0",
        )
    }
    captured: dict[str, object] = {}

    class _ConfiguredMiniMaxAnalyzer:
        def __init__(self, base_url: str, api_key: str, model: str) -> None:
            captured["credentials"] = (base_url, api_key, model)

    class _UnexpectedOpenAIAnalyzer:
        def __init__(self, *args: object, **kwargs: object) -> None:
            raise AssertionError("MiniMax provider must not use OpenAI-compatible analyzer")

    monkeypatch.setenv("OPPORTUNITY_RADAR_LLM_PROVIDER", "minimax")
    monkeypatch.setenv("OPPORTUNITY_RADAR_LLM_API_KEY", "runtime-secret")
    if base_url is None:
        monkeypatch.delenv("OPPORTUNITY_RADAR_LLM_BASE_URL", raising=False)
    else:
        monkeypatch.setenv("OPPORTUNITY_RADAR_LLM_BASE_URL", base_url)
    if model is None:
        monkeypatch.delenv("OPPORTUNITY_RADAR_LLM_MODEL", raising=False)
    else:
        monkeypatch.setenv("OPPORTUNITY_RADAR_LLM_MODEL", model)
    monkeypatch.setattr(cli, "load_sources", lambda path: configured)
    monkeypatch.setattr(
        cli,
        "load_compliance_sources",
        lambda path: {"miit": _verified_compliance_source("miit")},
    )
    monkeypatch.setattr(cli, "build_sources", lambda selected: {"miit": object()})
    monkeypatch.setattr(cli, "load_industry_codes", lambda path: {"C", "C34"})
    monkeypatch.setattr(cli, "MiniMaxAnthropicAnalyzer", _ConfiguredMiniMaxAnalyzer)
    monkeypatch.setattr(cli, "OpenAICompatibleAnalyzer", _UnexpectedOpenAIAnalyzer)
    monkeypatch.setattr(
        cli,
        "run_pipeline",
        lambda *args: (tmp_path / "result.xlsx", tmp_path / "report.json"),
    )

    cli.main(["run", "--sources", "miit"])

    assert captured["credentials"] == (
        expected_base_url,
        "runtime-secret",
        expected_model,
    )


def test_cli_rejects_explicitly_selected_disabled_source(
    monkeypatch: Any,
    capsys: pytest.CaptureFixture[str],
) -> None:
    from opportunity_radar import cli

    configured = {
        "disabled": SourceConfig(
            "disabled",
            "停用来源",
            "全国",
            ("https://disabled.example.gov.cn/list",),
            ("disabled.example.gov.cn",),
            enabled=False,
        )
    }
    monkeypatch.setattr(cli, "load_sources", lambda path: configured)
    monkeypatch.setattr(cli, "load_compliance_sources", lambda path: {})

    with pytest.raises(SystemExit):
        cli.main(["run", "--sources", "disabled"])

    assert "not eligible: disabled (not compliance-registered)" in capsys.readouterr().err


def test_cli_rejects_candidate_before_building_sources(
    monkeypatch: Any, capsys: pytest.CaptureFixture[str]
) -> None:
    from opportunity_radar import cli

    candidate = ComplianceSource(
        source_id="state_council_policy_library",
        display_name="国务院政策文件库",
        phase="candidate",
        enabled=False,
        terms="自动访问条款待确认。",
        terms_confirmed=None,
        registration="unknown",
        registration_completed=None,
        authorization="unknown",
        rate_limit=None,
        selected_data_scope="unknown",
        field_permission_confirmed=None,
        evidence_url="https://example.gov.cn/evidence",
        verified_at=None,
        review_due_at=date(2026, 10, 27),
        owner="unassigned",
        available_fields=("标题",),
    )
    monkeypatch.setattr(cli, "load_sources", lambda path: {})
    monkeypatch.setattr(
        cli, "load_compliance_sources", lambda path: {candidate.source_id: candidate}
    )
    monkeypatch.setattr(
        cli,
        "build_sources",
        lambda _: (_ for _ in ()).throw(AssertionError("no network")),
    )

    with pytest.raises(SystemExit):
        cli.main(["run", "--sources", "state_council_policy_library"])

    assert "not eligible: state_council_policy_library (phase=candidate)" in capsys.readouterr().err


def test_cli_default_selects_only_verified_enabled_registered_sources(
    monkeypatch: Any, tmp_path: Path
) -> None:
    from opportunity_radar import cli

    configured = {
        "verified": SourceConfig(
            "verified",
            "已核验",
            "全国",
            ("https://verified.example.gov.cn",),
            ("verified.example.gov.cn",),
            request_interval_seconds=5,
            adapter_version="1.0.0",
        ),
        "legacy": SourceConfig(
            "legacy",
            "未登记",
            "全国",
            ("https://legacy.example.gov.cn",),
            ("legacy.example.gov.cn",),
        ),
    }
    verified = _verified_compliance_source("verified")
    captured: dict[str, object] = {}

    def fake_build_sources(selected: dict[str, SourceConfig]) -> dict[str, object]:
        captured["selected"] = tuple(selected)
        return {"verified": object()}

    monkeypatch.setattr(cli, "load_sources", lambda path: configured)
    monkeypatch.setattr(cli, "load_compliance_sources", lambda path: {"verified": verified})
    monkeypatch.setattr(cli, "build_sources", fake_build_sources)
    monkeypatch.setattr(cli, "load_industry_codes", lambda path: {"C": "制造业", "C34": "通用设备制造业"})
    monkeypatch.setattr(cli, "OpenAICompatibleAnalyzer", lambda *args: object())
    monkeypatch.setattr(cli, "run_pipeline", lambda *args: (tmp_path / "result.xlsx", tmp_path / "report.json"))
    monkeypatch.setenv("OPPORTUNITY_RADAR_LLM_API_KEY", "test-key")
    monkeypatch.setenv("OPPORTUNITY_RADAR_LLM_MODEL", "test-model")

    cli.main(["run"])

    assert captured["selected"] == ("verified",)


@pytest.mark.parametrize(
    ("requested", "configured", "compliance_error", "compliance", "expected"),
    [
        (
            "missing",
            {},
            None,
            {},
            "missing (not compliance-registered)",
        ),
        (
            "legacy",
            {
                "legacy": SourceConfig(
                    "legacy",
                    "未登记来源",
                    "全国",
                    ("https://legacy.example.gov.cn/list",),
                    ("legacy.example.gov.cn",),
                )
            },
            None,
            {},
            "legacy (not compliance-registered)",
        ),
        (
            "candidate",
            {},
            None,
            {
                "candidate": replace(
                    _verified_compliance_source("candidate"),
                    phase="candidate",
                    enabled=False,
                    terms_confirmed=None,
                    registration="unknown",
                    registration_completed=None,
                    authorization="unknown",
                    rate_limit=None,
                    selected_data_scope="unknown",
                    field_permission_confirmed=None,
                    verified_at=None,
                    owner="unassigned",
                )
            },
            "candidate (phase=candidate)",
        ),
        (
            "expired",
            {
                "expired": SourceConfig(
                    "expired",
                    "过期来源",
                    "全国",
                    ("https://expired.example.gov.cn/list",),
                    ("expired.example.gov.cn",),
                    adapter_version="1.0.0",
                )
            },
            None,
            {
                "expired": replace(
                    _verified_compliance_source("expired"),
                    verified_at=date(2026, 5, 1),
                    review_due_at=date(2026, 7, 20),
                )
            },
            "expired (review_due_at=expired)",
        ),
        (
            "invalid",
            {},
            ValueError("source invalid: phase must be valid"),
            {},
            "source invalid: phase must be valid",
        ),
    ],
)
def test_cli_rejects_ineligible_source_before_all_runtime_side_effects(
    monkeypatch: Any,
    capsys: pytest.CaptureFixture[str],
    requested: str,
    configured: dict[str, SourceConfig],
    compliance_error: ValueError | None,
    compliance: dict[str, ComplianceSource],
    expected: str,
) -> None:
    from opportunity_radar import cli

    monkeypatch.setattr(cli, "load_sources", lambda path: configured)
    if compliance_error is None:
        monkeypatch.setattr(cli, "load_compliance_sources", lambda path: compliance)
    else:
        def fail_compliance_load(path: Path) -> dict[str, ComplianceSource]:
            raise compliance_error

        monkeypatch.setattr(cli, "load_compliance_sources", fail_compliance_load)
    monkeypatch.setattr(
        cli,
        "build_sources",
        lambda configs: (_ for _ in ()).throw(
            AssertionError("must not construct source adapters")
        ),
    )
    monkeypatch.setattr(
        cli.os,
        "getenv",
        lambda *args: (_ for _ in ()).throw(
            AssertionError("must not read LLM environment")
        ),
    )
    monkeypatch.setattr(
        cli,
        "DocumentRetriever",
        lambda: (_ for _ in ()).throw(
            AssertionError("must not construct the retriever")
        ),
    )

    with pytest.raises(SystemExit):
        cli.main(["run", "--sources", requested])

    assert expected in capsys.readouterr().err


def test_cli_rejects_placeholder_adapter_version_before_runtime_side_effects(
    monkeypatch: Any,
    capsys: pytest.CaptureFixture[str],
) -> None:
    from opportunity_radar import cli

    source = SourceConfig(
        "verified",
        "已核验来源",
        "全国",
        ("https://verified.example.gov.cn/list",),
        ("verified.example.gov.cn",),
        request_interval_seconds=5,
    )
    monkeypatch.setattr(cli, "load_sources", lambda path: {"verified": source})
    monkeypatch.setattr(
        cli,
        "load_compliance_sources",
        lambda path: {"verified": _verified_compliance_source("verified")},
    )
    monkeypatch.setattr(
        cli,
        "build_sources",
        lambda configs: (_ for _ in ()).throw(
            AssertionError("must not construct source adapters")
        ),
    )

    with pytest.raises(SystemExit):
        cli.main(["run", "--sources", "verified"])

    assert "verified (adapter_version=unregistered)" in capsys.readouterr().err


@pytest.mark.parametrize(
    ("interval", "expected_reason"),
    [
        (0, "request_interval_seconds must be a finite positive number"),
        (float("nan"), "request_interval_seconds must be a finite positive number"),
        (float("inf"), "request_interval_seconds must be a finite positive number"),
        (4.99, "request_interval_seconds=4.99 below required=5"),
    ],
)
def test_cli_rejects_incompatible_verified_pacing_before_runtime_side_effects(
    monkeypatch: Any,
    capsys: pytest.CaptureFixture[str],
    interval: float,
    expected_reason: str,
) -> None:
    from opportunity_radar import cli

    source = SourceConfig(
        "verified",
        "已核验来源",
        "全国",
        ("https://verified.example.gov.cn/list",),
        ("verified.example.gov.cn",),
        request_interval_seconds=interval,
        adapter_version="1.0.0",
    )
    monkeypatch.setattr(cli, "load_sources", lambda path: {"verified": source})
    monkeypatch.setattr(
        cli,
        "load_compliance_sources",
        lambda path: {"verified": _verified_compliance_source("verified")},
    )
    monkeypatch.setattr(
        cli,
        "build_sources",
        lambda configs: (_ for _ in ()).throw(
            AssertionError("must not construct source adapters")
        ),
    )
    monkeypatch.setattr(
        cli.os,
        "getenv",
        lambda *args: (_ for _ in ()).throw(
            AssertionError("must not read LLM environment")
        ),
    )
    monkeypatch.setattr(
        cli,
        "DocumentRetriever",
        lambda: (_ for _ in ()).throw(
            AssertionError("must not construct the retriever")
        ),
    )

    with pytest.raises(SystemExit):
        cli.main(["run", "--sources", "verified"])

    assert expected_reason in capsys.readouterr().err


def test_compliant_selection_accepts_adapter_interval_at_required_boundary() -> None:
    from opportunity_radar.cli import _select_compliant_sources

    source = SourceConfig(
        "verified",
        "已核验来源",
        "全国",
        ("https://verified.example.gov.cn/list",),
        ("verified.example.gov.cn",),
        request_interval_seconds=5,
        adapter_version="1.0.0",
    )

    selected, audit = _select_compliant_sources(
        ("verified",),
        {"verified": source},
        {"verified": _verified_compliance_source("verified")},
        date(2026, 7, 29),
    )

    assert selected == ("verified",)
    assert audit[0].adapter_version == "1.0.0"
