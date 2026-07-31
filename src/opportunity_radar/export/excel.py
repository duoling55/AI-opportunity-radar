from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SHEET_NAMES = ("重点商机", "政策观察")
HEADERS = [
    "商机等级",
    "商机评分",
    "政策名称",
    "政策摘要",
    "政策层级",
    "适用地区",
    "发布机构",
    "政策文号",
    "发布日期",
    "申报开始日期",
    "申报截止日期",
    "支持方向",
    "适用企业条件",
    "国标行业门类名称",
    "国标行业门类代码",
    "国标行业大类名称",
    "国标行业大类代码",
    "业务行业标签",
    "行业判断置信度",
    "机会场景",
    "融资租赁关联度",
    "推荐理由",
    "评分理由",
    "政策原文依据",
    "依据位置",
    "推荐动作",
    "推荐联系时间",
    "行业营销开场白",
    "风险与限制",
    "复核原因",
    "政策原文链接",
    "附件链接",
    "数据来源",
    "采集时间",
    "AI 分析时间",
    "免责声明",
]
LINK_HEADERS = ("政策原文链接", "附件链接")
DATE_HEADERS = ("发布日期", "申报开始日期", "申报截止日期")
DATETIME_HEADERS = ("采集时间", "AI 分析时间")
LONG_TEXT_HEADERS = {
    "政策摘要",
    "支持方向",
    "适用企业条件",
    "机会场景",
    "推荐理由",
    "评分理由",
    "政策原文依据",
    "依据位置",
    "推荐动作",
    "行业营销开场白",
    "风险与限制",
    "复核原因",
    "政策原文链接",
    "附件链接",
    "免责声明",
}
GRADE_FILLS = {
    "A": "C6EFCE",
    "B": "D9EAF7",
    "C": "FFF2CC",
    "观察": "E7E6E6",
}


@dataclass(frozen=True)
class ExportRow:
    sheet_name: str
    values: dict[str, str | int | float | date | datetime]


def _available_path(output_dir: Path, run_date: date) -> Path:
    base_name = f"policy-opportunities-{run_date.isoformat()}"
    candidate = output_dir / f"{base_name}.xlsx"
    sequence = 1
    while candidate.exists():
        candidate = output_dir / f"{base_name}-{sequence}.xlsx"
        sequence += 1
    return candidate


def _write_sheet(workbook: Workbook, sheet_name: str, rows: list[ExportRow]) -> None:
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    sheet.freeze_panes = "A2"

    for export_row in sorted(rows, key=_sort_key, reverse=True):
        sheet.append([export_row.values.get(header, "") for header in HEADERS])
        grade = str(export_row.values.get("商机等级", "")).strip()
        if fill_color := GRADE_FILLS.get(grade):
            sheet.cell(sheet.max_row, 1).fill = PatternFill("solid", fgColor=fill_color)
        for header in LINK_HEADERS:
            links = _valid_links(export_row.values.get(header, ""))
            if links:
                cell = sheet.cell(sheet.max_row, HEADERS.index(header) + 1)
                cell.value = "\n".join(links)
                cell.hyperlink = links[0]
                cell.style = "Hyperlink"
        for header in DATE_HEADERS:
            sheet.cell(sheet.max_row, HEADERS.index(header) + 1).number_format = "yyyy-mm-dd"
        for header in DATETIME_HEADERS:
            sheet.cell(sheet.max_row, HEADERS.index(header) + 1).number_format = (
                "yyyy-mm-dd hh:mm:ss"
            )
        for header in HEADERS:
            sheet.cell(sheet.max_row, HEADERS.index(header) + 1).alignment = Alignment(
                vertical="top",
                wrap_text=header in LONG_TEXT_HEADERS,
            )

    sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{sheet.max_row}"
    for column_index, header in enumerate(HEADERS, start=1):
        width = 12
        if header in LONG_TEXT_HEADERS:
            width = 42
        elif header in {"政策名称", "发布机构", "数据来源"}:
            width = 28
        elif header in LINK_HEADERS:
            width = 36
        elif header in DATE_HEADERS:
            width = 14
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def _sort_key(row: ExportRow) -> tuple[float, int]:
    score_value = row.values.get("商机评分", 0)
    try:
        score = float(score_value)
    except (TypeError, ValueError):
        score = 0
    published = row.values.get("发布日期")
    if isinstance(published, datetime):
        ordinal = published.date().toordinal()
    elif isinstance(published, date):
        ordinal = published.toordinal()
    else:
        try:
            ordinal = date.fromisoformat(str(published)).toordinal()
        except ValueError:
            ordinal = 0
    return score, ordinal


def _valid_links(value: object) -> list[str]:
    links = [
        item.strip()
        for item in re.split(r"[、\n;；]+", str(value))
        if item.strip()
    ]
    return [
        link
        for link in links
        if urlparse(link).scheme in {"http", "https"} and urlparse(link).netloc
    ]


def export_workbook(rows: list[ExportRow], output_dir: Path, run_date: date) -> Path:
    """Create a two-sheet business workbook without replacing an earlier export."""
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = _available_path(output_dir, run_date)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in SHEET_NAMES:
        _write_sheet(workbook, sheet_name, [row for row in rows if row.sheet_name == sheet_name])
    workbook.save(output_path)
    return output_path
