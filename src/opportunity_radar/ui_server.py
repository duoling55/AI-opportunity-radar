from __future__ import annotations

import argparse
import json
import mimetypes
import os
import re
import subprocess
import sys
import threading
import webbrowser
from dataclasses import dataclass
from datetime import UTC, date, datetime
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook

from opportunity_radar.analysis.prompts import (
    SYSTEM_PROMPT,
    USER_PROMPT_PLACEHOLDERS,
    USER_PROMPT_TEMPLATE,
    validate_user_prompt_template,
)
from opportunity_radar.collection import filter_collectable, load_batch
from opportunity_radar.config import SourceConfig, load_sources
from opportunity_radar.diagnostics import safe_url
from opportunity_radar.discovery.service import DiscoveryService
from opportunity_radar.sources.registry import SOURCE_TYPES

PROJECT_ROOT = Path(__file__).resolve().parents[2]
STATIC_DIRECTORY = Path(__file__).with_name("ui_static")
SOURCE_CONFIG_PATH = PROJECT_ROOT / "config" / "sources.json"
PROMPT_CONFIG_PATH = PROJECT_ROOT / "config" / "analysis_prompts.json"
COMPLIANCE_CONFIG_PATH = PROJECT_ROOT / "config" / "compliance_sources.json"
DISCOVERY_PORTALS_PATH = PROJECT_ROOT / "config" / "discovery_portals.json"
DISCOVERY_KEYWORDS_PATH = PROJECT_ROOT / "config" / "discovery_keywords.json"
DISCOVERY_REPORT_DIRECTORY = PROJECT_ROOT / "data" / "discovery"
BATCH_DIRECTORY = PROJECT_ROOT / "data" / "normalized" / "batches"
RAW_DIRECTORY = PROJECT_ROOT / "data" / "raw"
OUTPUT_DIRECTORY = PROJECT_ROOT / "outputs"
JOB_DIRECTORY = PROJECT_ROOT / "data" / "ui_jobs"
MAX_REQUEST_BYTES = 1024 * 1024


@dataclass
class UiJob:
    job_id: str
    label: str
    process: subprocess.Popen[str]
    log_path: Path
    started_at: datetime
    pipe_path: Path | None = None  # 用于发送控制消息的命名管道


_JOBS: list[UiJob] = []
_JOBS_LOCK = threading.Lock()


def _json_default(value: object) -> str:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def _files(directory: Path, pattern: str) -> list[Path]:
    if not directory.exists():
        return []
    return sorted(directory.glob(pattern), key=lambda item: item.stat().st_mtime, reverse=True)


def _file_item(path: Path) -> dict[str, object]:
    return {
        "name": path.name,
        "size": path.stat().st_size,
        "modified_at": datetime.fromtimestamp(path.stat().st_mtime, UTC).isoformat(),
    }


def _named_file(directory: Path, name: str, pattern: str) -> Path:
    if not name or Path(name).name != name:
        raise ValueError("非法文件名")
    candidates = {path.name: path for path in _files(directory, pattern)}
    if name not in candidates:
        raise FileNotFoundError(name)
    return candidates[name]


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _source_payload() -> list[dict[str, object]]:
    return [
        {
            "source_id": source.source_id,
            "display_name": source.display_name,
            "region": source.region,
            "list_urls": list(source.list_urls),
            "allowed_domains": list(source.allowed_domains),
            "enabled": source.enabled,
            "request_interval_seconds": source.request_interval_seconds,
            "adapter_version": source.adapter_version,
            "origin": source.origin,
        }
        for source in load_sources(SOURCE_CONFIG_PATH).values()
    ]


def _validate_sources(items: object) -> list[dict[str, object]]:
    if not isinstance(items, list):
        raise TypeError("信源配置必须是列表")
    original_ids = set(load_sources(SOURCE_CONFIG_PATH))
    payload: list[dict[str, object]] = []
    seen: set[str] = set()
    for item in items:
        if not isinstance(item, dict):
            raise TypeError("信源配置项格式错误")
        source_id = str(item.get("source_id", "")).strip()
        if source_id in seen:
            raise ValueError(f"信源 ID 重复：{source_id}")
        origin = str(item.get("origin", "manual")).strip()
        adapter_version = str(item.get("adapter_version", "")).strip()
        is_generic_discovery = origin == "discovery" and adapter_version == "generic"
        if source_id not in original_ids:
            if not is_generic_discovery:
                raise ValueError(
                    f"仅允许新增 origin=discovery 且 generic 适配器的信源：{source_id}"
                )
        elif source_id not in SOURCE_TYPES and not is_generic_discovery:
            raise ValueError(f"信源 ID 不可修改，且必须有现成适配器：{source_id}")
        seen.add(source_id)
        list_urls = tuple(
            str(value).strip()
            for value in item.get("list_urls", [])
            if str(value).strip()
        )
        allowed_domains = tuple(
            str(value).strip()
            for value in item.get("allowed_domains", [])
            if str(value).strip()
        )
        if not list_urls or not allowed_domains:
            raise ValueError(f"{source_id} 必须配置列表网址和允许域名")
        for url in list_urls:
            parsed = urlparse(url)
            if parsed.scheme != "https" or not parsed.hostname:
                raise ValueError(f"{source_id} 列表网址必须是有效 HTTPS 地址：{url}")
            if parsed.hostname not in allowed_domains:
                raise ValueError(f"{source_id} 的网址域名未列入允许域名：{parsed.hostname}")
        interval = float(item.get("request_interval_seconds", 0))
        if interval < 0:
            raise ValueError(f"{source_id} 的请求间隔不能小于 0")
        source = SourceConfig(
            source_id=source_id,
            display_name=str(item.get("display_name", "")).strip(),
            region=str(item.get("region", "")).strip(),
            list_urls=list_urls,
            allowed_domains=allowed_domains,
            enabled=bool(item.get("enabled", False)),
            request_interval_seconds=interval,
            adapter_version=adapter_version,
            origin=origin,
        )
        if not source.display_name or not source.region:
            raise ValueError(f"{source_id} 的名称和地区不能为空")
        payload.append(
            {
                "source_id": source.source_id,
                "display_name": source.display_name,
                "region": source.region,
                "list_urls": list(source.list_urls),
                "allowed_domains": list(source.allowed_domains),
                "enabled": source.enabled,
                "request_interval_seconds": source.request_interval_seconds,
                "adapter_version": source.adapter_version,
                "origin": source.origin,
            }
        )
    if not original_ids.issubset(seen):
        raise ValueError("不能在此页面删除已有信源")
    return payload


def _save_sources(items: object) -> list[dict[str, object]]:
    payload = _validate_sources(items)
    temporary = SOURCE_CONFIG_PATH.with_suffix(".json.tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    temporary.replace(SOURCE_CONFIG_PATH)
    load_sources(SOURCE_CONFIG_PATH)
    return payload


def _prompt_payload() -> dict[str, object]:
    payload: dict[str, object] = {}
    if PROMPT_CONFIG_PATH.exists():
        payload = _read_json(PROMPT_CONFIG_PATH)
    system_prompt = str(payload.get("system_prompt", SYSTEM_PROMPT)).strip() or SYSTEM_PROMPT
    user_prompt_template = validate_user_prompt_template(
        str(payload.get("user_prompt_template", USER_PROMPT_TEMPLATE))
    )
    return {
        "system_prompt": system_prompt,
        "user_prompt_template": user_prompt_template,
        "placeholders": list(USER_PROMPT_PLACEHOLDERS),
        "built_in_system_prompt": SYSTEM_PROMPT,
        "built_in_user_prompt_template": USER_PROMPT_TEMPLATE,
    }


def _save_prompts(payload: dict[str, Any]) -> dict[str, object]:
    system_prompt = str(payload.get("system_prompt", "")).strip()
    if not system_prompt:
        raise ValueError("系统提示词不能为空")
    user_prompt_template = validate_user_prompt_template(
        str(payload.get("user_prompt_template", ""))
    )
    saved = {
        "system_prompt": system_prompt,
        "user_prompt_template": user_prompt_template,
    }
    temporary = PROMPT_CONFIG_PATH.with_suffix(".json.tmp")
    temporary.write_text(
        json.dumps(saved, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    temporary.replace(PROMPT_CONFIG_PATH)
    return _prompt_payload()


def _summary_payload() -> dict[str, object]:
    batches = _files(BATCH_DIRECTORY, "policy-batch-*.json")
    workbooks = _files(OUTPUT_DIRECTORY, "policy-opportunities-*.xlsx")
    reports = _files(OUTPUT_DIRECTORY, "policy-opportunities-*-report*.json")
    latest_documents = 0
    latest_batch = None
    if batches:
        batch = load_batch(batches[0])
        latest_documents = len(batch.documents)
        latest_batch = batches[0].name
    latest_report = _read_json(reports[0]) if reports else {}
    with _JOBS_LOCK:
        active_jobs = sum(job.process.poll() is None for job in _JOBS)
    return {
        "source_count": len(load_sources(SOURCE_CONFIG_PATH)),
        "latest_documents": latest_documents,
        "latest_batch": latest_batch,
        "workbook_count": len(workbooks),
        "active_jobs": active_jobs,
        "latest_report": latest_report,
    }


def _batch_list_payload() -> list[dict[str, object]]:
    result: list[dict[str, object]] = []
    for path in _files(BATCH_DIRECTORY, "policy-batch-*.json"):
        batch = load_batch(path)
        result.append(
            {
                **_file_item(path),
                "start_date": batch.start_date,
                "end_date": batch.end_date,
                "document_count": len(batch.documents),
                "development_mode": batch.development_mode,
                "report": batch.report.__dict__,
            }
        )
    return result


def _batch_payload(name: str, query: str = "", source_id: str = "") -> dict[str, object]:
    path = _named_file(BATCH_DIRECTORY, name, "policy-batch-*.json")
    batch = load_batch(path)
    normalized_query = query.casefold().strip()
    documents: list[dict[str, object]] = []
    for document in batch.documents:
        searchable = " ".join(
            [
                document.title,
                document.publisher or "",
                document.document_number or "",
                document.normalized_text,
            ]
        ).casefold()
        if source_id and document.source_id != source_id:
            continue
        if normalized_query and normalized_query not in searchable:
            continue
        documents.append(
            {
                "policy_id": document.policy_id,
                "title": document.title,
                "source_id": document.source_id,
                "source_name": document.source_name,
                "region": document.region,
                "publisher": document.publisher,
                "document_number": document.document_number,
                "publish_date": document.publish_date,
                "text_length": len(document.normalized_text),
                "attachment_count": len(document.attachment_urls),
                "detail_url": str(document.detail_url),
            }
        )
    return {
        "name": path.name,
        "start_date": batch.start_date,
        "end_date": batch.end_date,
        "development_mode": batch.development_mode,
        "source_ids": list(batch.source_ids),
        "report": batch.report.__dict__,
        "documents": documents,
    }


def _document_payload(batch_name: str, policy_id: str) -> dict[str, object]:
    path = _named_file(BATCH_DIRECTORY, batch_name, "policy-batch-*.json")
    batch = load_batch(path)
    for document in batch.documents:
        if document.policy_id == policy_id:
            return document.model_dump(mode="json")
    raise FileNotFoundError(policy_id)


def _selection_batch_path(batch_name: str, policy_ids: object) -> tuple[Path, int]:
    if not isinstance(policy_ids, list) or not policy_ids:
        raise ValueError("请至少选择一篇公文进行分析")
    if any(not isinstance(policy_id, str) or not policy_id.strip() for policy_id in policy_ids):
        raise TypeError("所选公文 ID 格式错误")
    requested = list(dict.fromkeys(policy_ids))
    source_path = _named_file(BATCH_DIRECTORY, batch_name, "policy-batch-*.json")
    payload = _read_json(source_path)
    documents = payload.get("documents", [])
    if not isinstance(documents, list):
        raise TypeError("原始批次文档格式错误")
    selected = [
        document
        for document in documents
        if isinstance(document, dict) and document.get("policy_id") in requested
    ]
    selected_ids = {str(document["policy_id"]) for document in selected}
    missing = [policy_id for policy_id in requested if policy_id not in selected_ids]
    if missing:
        raise ValueError("所选公文不在当前批次：" + "、".join(missing))
    selection_directory = JOB_DIRECTORY / "analysis_batches"
    selection_directory.mkdir(parents=True, exist_ok=True)
    selection_id = datetime.now(UTC).strftime("%Y%m%d-%H%M%S-%f")
    selection_path = selection_directory / f"policy-selection-{selection_id}.json"
    payload["created_at"] = datetime.now(UTC).isoformat()
    payload["documents"] = selected
    payload["report"] = {
        "discovered": len(selected),
        "collected": len(selected),
        "skipped": 0,
        "source_failures": 0,
        "parse_failures": 0,
    }
    payload["selection"] = {
        "source_batch": batch_name,
        "policy_ids": requested,
    }
    selection_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    load_batch(selection_path)
    return selection_path, len(selected)


def _start_job(label: str, arguments: list[str], environment: dict[str, str]) -> UiJob:
    JOB_DIRECTORY.mkdir(parents=True, exist_ok=True)
    job_id = datetime.now(UTC).strftime("%Y%m%d-%H%M%S-%f")
    log_path = JOB_DIRECTORY / f"{job_id}.log"
    pipe_dir = JOB_DIRECTORY / "pipes"
    pipe_dir.mkdir(parents=True, exist_ok=True)
    pipe_path = pipe_dir / f"{job_id}.pipe"

    creation_flags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
    environment = environment.copy()
    environment["PYTHONUNBUFFERED"] = "1"
    environment["OPPORTUNITY_RADAR_JOB_PIPE"] = str(pipe_path)  # 传递给子进程用于监听停止消息

    log_handle = log_path.open("w", encoding="utf-8")
    try:
        safe_metadata = {
            "job_id": job_id,
            "label": label,
            "command": arguments[0] if arguments else "",
            "provider": environment.get("OPPORTUNITY_RADAR_LLM_PROVIDER", ""),
            "model": environment.get("OPPORTUNITY_RADAR_LLM_MODEL", ""),
            "base_url": safe_url(environment.get("OPPORTUNITY_RADAR_LLM_BASE_URL", "")),
        }
        log_handle.write(
            f"{datetime.now(UTC).isoformat()} INFO opportunity_radar.ui - "
            f"页面任务启动 {json.dumps(safe_metadata, ensure_ascii=False)}\n"
        )
        log_handle.flush()
        process = subprocess.Popen(
            [sys.executable, "-u", "-m", "opportunity_radar", *arguments],
            cwd=PROJECT_ROOT,
            env=environment,
            stdout=log_handle,
            stderr=subprocess.STDOUT,
            text=True,
            creationflags=creation_flags,
        )
    finally:
        log_handle.close()
    job = UiJob(job_id, label, process, log_path, datetime.now(UTC), pipe_path)
    with _JOBS_LOCK:
        _JOBS.append(job)
    return job


def _stop_job(job_id: str) -> dict[str, object]:
    """停止指定的采集或分析任务，通过消息通信优雅停止。

    优先通过命名管道发送 STOP 消息，让子进程优雅关闭浏览器和资源；
    如果 5 秒内未退出，则强制终止进程树。
    """
    with _JOBS_LOCK:
        job = next((j for j in _JOBS if j.job_id == job_id), None)
        if not job:
            raise FileNotFoundError(f"任务不存在：{job_id}")
        if job.process.poll() is not None:
            return {"job_id": job_id, "status": "already_stopped", "return_code": job.process.poll()}

        # 尝试通过管道发送优雅停止消息
        stopped_gracefully = False
        if job.pipe_path and job.pipe_path.exists():
            try:
                # 非阻塞写入管道发送停止消息
                pipe_fd = os.open(str(job.pipe_path), os.O_WRONLY | os.O_NONBLOCK)
                try:
                    os.write(pipe_fd, b"STOP\n")
                    stopped_gracefully = True
                finally:
                    os.close(pipe_fd)
            except (OSError, IOError) as e:
                LOGGER.warning("管道消息发送失败：%s", e)

        # 等待最多 5 秒让进程优雅退出
        if stopped_gracefully:
            try:
                return_code = job.process.wait(timeout=5)
                with _JOBS_LOCK:
                    if job in _JOBS:
                        _JOBS.remove(job)
                return {"job_id": job_id, "status": "stopped_gracefully", "return_code": return_code}
            except subprocess.TimeoutExpired:
                pass  # 超时后继续强制终止

        # 强制终止进程树（包括 Playwright 浏览器子进程）
        try:
            if os.name == "nt":
                # Windows: 使用 taskkill 终止进程树
                subprocess.run(
                    ["taskkill", "/F", "/T", "/PID", str(job.process.pid)],
                    capture_output=True,
                    timeout=5,
                )
            else:
                # Unix: 发送 SIGTERM 然后 SIGKILL
                import signal
                job.process.terminate()
                try:
                    job.process.wait(timeout=3)
                except subprocess.TimeoutExpired:
                    job.process.kill()
                    job.process.wait(timeout=3)
        except Exception as e:
            LOGGER.warning("停止任务时遇到异常：%s", e)

        # 从活动列表中移除
        with _JOBS_LOCK:
            if job in _JOBS:
                _JOBS.remove(job)
        return {"job_id": job_id, "status": "force_stopped", "return_code": job.process.poll()}


def _report_from_job_log(log: str) -> dict[str, Any] | None:
    matches = re.findall(r"^Report:\s*(.+?)\s*$", log, flags=re.MULTILINE)
    if not matches:
        return None
    report_path = Path(matches[-1])
    if not report_path.is_absolute():
        report_path = PROJECT_ROOT / report_path
    try:
        resolved = report_path.resolve()
        if OUTPUT_DIRECTORY.resolve() not in resolved.parents:
            return None
        return _read_json(resolved)
    except (json.JSONDecodeError, OSError, TypeError, ValueError):
        return None


def _job_payload(job: UiJob) -> dict[str, object]:
    return_code = job.process.poll()
    full_log = (
        job.log_path.read_text(encoding="utf-8", errors="replace")
        if job.log_path.exists()
        else ""
    )
    report = _report_from_job_log(full_log)
    has_partial_failures = bool(
        report
        and any(
            int(report.get(field, 0))
            for field in ("source_failures", "parse_failures", "analysis_failures")
        )
    )
    if return_code is None:
        status = "running"
    elif return_code != 0:
        status = "failed"
    elif has_partial_failures:
        status = "warning"
    else:
        status = "success"
    return {
        "job_id": job.job_id,
        "label": job.label,
        "status": status,
        "return_code": return_code,
        "started_at": job.started_at,
        "log": full_log[-20000:],
        "log_size": len(full_log.encode("utf-8")),
        "log_url": f"/download/job-log?id={job.job_id}",
        "report": report,
    }


def _jobs_payload() -> list[dict[str, object]]:
    with _JOBS_LOCK:
        return [_job_payload(job) for job in reversed(_JOBS[-12:])]


def _start_collection(payload: dict[str, Any]) -> dict[str, object]:
    configured = load_sources(SOURCE_CONFIG_PATH)
    source_ids = payload.get("source_ids", [])
    if not isinstance(source_ids, list) or not source_ids:
        raise ValueError("请至少选择一个信源")
    invalid = [
        source_id
        for source_id in source_ids
        if source_id not in configured or not configured[source_id].enabled
    ]
    if invalid:
        raise ValueError("信源不存在或已停用：" + "、".join(invalid))
    # 采集门控：discovery 信源须 verified AND enabled 才可采集
    selectable_ids = {
        str(item["source_id"])
        for item in filter_collectable(
            [
                {
                    "source_id": source_id,
                    "origin": getattr(configured[source_id], "origin", "manual"),
                }
                for source_id in source_ids
            ],
            compliance_path=str(COMPLIANCE_CONFIG_PATH),
        )
    }
    blocked = [source_id for source_id in source_ids if source_id not in selectable_ids]
    if blocked:
        raise ValueError("信源未核验，不可采集：" + "、".join(blocked))
    start_date = date.fromisoformat(str(payload.get("start_date", "")))
    end_date = date.fromisoformat(str(payload.get("end_date", "")))
    if start_date > end_date:
        raise ValueError("开始日期不能晚于结束日期")
    browser_mode = str(payload.get("browser_mode", "always"))
    if browser_mode not in {"always", "fallback", "off"}:
        raise ValueError("浏览器模式无效")
    max_pages = int(payload.get("max_pages", 20))
    if not 1 <= max_pages <= 100:
        raise ValueError("最多列表页必须在 1 到 100 之间")
    arguments = [
        "collect",
        "--start-date",
        start_date.isoformat(),
        "--end-date",
        end_date.isoformat(),
        "--sources",
        ",".join(source_ids),
        "--browser",
        browser_mode,
        "--max-pages",
        str(max_pages),
    ]
    if bool(payload.get("headed", False)):
        arguments.append("--headed")
    if bool(payload.get("development_mode", True)):
        arguments.append("--dev-unverified-sources")
    job = _start_job(f"采集 {len(source_ids)} 个信源", arguments, os.environ.copy())
    return _job_payload(job)


def _start_analysis(payload: dict[str, Any]) -> dict[str, object]:
    batch_name = str(payload.get("batch_name", ""))
    batch_path = _named_file(BATCH_DIRECTORY, batch_name, "policy-batch-*.json")
    selected_count: int | None = None
    if "policy_ids" in payload:
        batch_path, selected_count = _selection_batch_path(
            batch_name,
            payload.get("policy_ids"),
        )
    provider = str(payload.get("provider", "openai")).strip().lower()
    if provider not in {"openai", "minimax"}:
        raise ValueError("API 协议无效")
    base_url = str(payload.get("base_url", "")).strip()
    model = str(payload.get("model", "")).strip()
    api_key = str(payload.get("api_key", "")).strip() or os.getenv(
        "OPPORTUNITY_RADAR_LLM_API_KEY", ""
    )
    if not api_key:
        raise ValueError("请输入 API Key，或在启动后台前设置环境变量")
    if not base_url or not model:
        raise ValueError("Base URL 和模型不能为空")
    configured_prompts = _prompt_payload()
    system_prompt = str(
        payload.get("system_prompt", configured_prompts["system_prompt"])
    ).strip()
    if not system_prompt:
        raise ValueError("系统提示词不能为空")
    user_prompt_template = validate_user_prompt_template(
        str(
            payload.get(
                "user_prompt_template",
                configured_prompts["user_prompt_template"],
            )
        )
    )
    environment = os.environ.copy()
    environment.update(
        {
            "OPPORTUNITY_RADAR_LLM_PROVIDER": provider,
            "OPPORTUNITY_RADAR_LLM_API_KEY": api_key,
            "OPPORTUNITY_RADAR_LLM_BASE_URL": base_url,
            "OPPORTUNITY_RADAR_LLM_MODEL": model,
            "OPPORTUNITY_RADAR_SYSTEM_PROMPT": system_prompt,
            "OPPORTUNITY_RADAR_USER_PROMPT_TEMPLATE": user_prompt_template,
        }
    )
    arguments = ["analyze-local", "--batch", str(batch_path)]
    if bool(payload.get("force", True)):
        arguments.append("--force")
    selection_label = f"（{selected_count} 篇）" if selected_count is not None else ""
    job = _start_job(f"分析 {batch_name}{selection_label}", arguments, environment)
    return _job_payload(job)


def _results_payload() -> dict[str, object]:
    return {
        "workbooks": [
            _file_item(path) for path in _files(OUTPUT_DIRECTORY, "policy-opportunities-*.xlsx")
        ],
        "reports": [
            {**_file_item(path), "data": _read_json(path)}
            for path in _files(OUTPUT_DIRECTORY, "policy-opportunities-*-report*.json")
        ],
    }


def _discovery_service() -> DiscoveryService:
    return DiscoveryService(
        compliance_path=str(COMPLIANCE_CONFIG_PATH),
        sources_path=str(SOURCE_CONFIG_PATH),
    )


def _discovery_candidates_payload() -> list[dict[str, object]]:
    return _discovery_service().list_candidates()


def _discovery_candidate_payload(source_id: str) -> dict[str, object]:
    return _discovery_service().get_candidate(source_id) or {}


def _discovery_portals_payload() -> list[dict[str, object]]:
    return _read_json(DISCOVERY_PORTALS_PATH) if DISCOVERY_PORTALS_PATH.exists() else []


def _discovery_keywords_payload() -> list[str]:
    items = _read_json(DISCOVERY_KEYWORDS_PATH) if DISCOVERY_KEYWORDS_PATH.exists() else []
    tags: list[str] = []
    for item in items:
        tag = item.get("tag")
        if tag and tag not in tags:
            tags.append(tag)
    return tags


def _discovery_reports_payload() -> list[dict[str, object]]:
    return [
        {**_file_item(path), "data": _read_json(path)}
        for path in _files(DISCOVERY_REPORT_DIRECTORY, "*-report.json")
    ]


def _start_discovery_search(payload: dict[str, Any]) -> dict[str, object]:
    keywords = str(payload.get("keywords", "all"))
    portals = str(payload.get("portals", "all"))
    arguments = ["search-sources", "--keywords", keywords, "--portals", portals]
    job = _start_job("信源搜索", arguments, os.environ.copy())
    return _job_payload(job)


def _discovery_promote(source_id: str, payload: dict[str, Any]) -> dict[str, object]:
    return _discovery_service().promote(
        source_id,
        reviewer=payload.get("reviewer", ""),
        override_not_recommended=bool(payload.get("override_not_recommended", False)),
    )


def _discovery_review(source_id: str, payload: dict[str, Any]) -> dict[str, object]:
    return _discovery_service().review(
        source_id,
        action=payload.get("action", ""),
        reason=payload.get("reason"),
        reviewer=payload.get("reviewer", ""),
        comment=payload.get("comment", ""),
    )


def _cell_value(value: object) -> object:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _workbook_payload(name: str, sheet_name: str = "", query: str = "") -> dict[str, object]:
    path = _named_file(OUTPUT_DIRECTORY, name, "policy-opportunities-*.xlsx")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = workbook.sheetnames
        selected_sheet = sheet_name if sheet_name in sheets else sheets[0]
        rows = workbook[selected_sheet].iter_rows(values_only=True)
        headers = [str(value or "") for value in next(rows, ())]
        normalized_query = query.casefold().strip()
        records: list[dict[str, object]] = []
        for values in rows:
            record = {
                header: _cell_value(value)
                for header, value in zip(headers, values, strict=False)
            }
            if normalized_query and normalized_query not in " ".join(
                str(value or "") for value in record.values()
            ).casefold():
                continue
            records.append(record)
        return {
            "name": path.name,
            "sheets": sheets,
            "selected_sheet": selected_sheet,
            "headers": headers,
            "rows": records,
        }
    finally:
        workbook.close()


class RadarRequestHandler(BaseHTTPRequestHandler):
    server_version = "OpportunityRadarUI/0.1"

    def log_message(self, format_string: str, *args: object) -> None:
        print(f"[ui] {self.address_string()} {format_string % args}")

    def _send_json(self, payload: object, status: HTTPStatus = HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=False, default=_json_default).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def _send_error_json(self, error: Exception, status: HTTPStatus) -> None:
        self._send_json({"error": str(error) or type(error).__name__}, status)

    def _read_body(self) -> dict[str, Any]:
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0 or length > MAX_REQUEST_BYTES:
            raise ValueError("请求内容为空或过大")
        payload = json.loads(self.rfile.read(length).decode("utf-8"))
        if not isinstance(payload, dict):
            raise TypeError("请求内容必须是 JSON 对象")
        return payload

    def _send_file(self, path: Path, *, attachment: bool = False) -> None:
        body = path.read_bytes()
        content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        if attachment:
            self.send_header(
                "Content-Disposition",
                f"attachment; filename*=UTF-8''{path.name}",
            )
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)
        try:
            if parsed.path == "/":
                self._send_file(STATIC_DIRECTORY / "index.html")
            elif parsed.path.startswith("/static/"):
                name = parsed.path.removeprefix("/static/")
                if Path(name).name != name:
                    raise FileNotFoundError(name)
                self._send_file(STATIC_DIRECTORY / name)
            elif parsed.path == "/api/summary":
                self._send_json(_summary_payload())
            elif parsed.path == "/api/sources":
                self._send_json(_source_payload())
            elif parsed.path == "/api/prompts":
                self._send_json(_prompt_payload())
            elif parsed.path == "/api/batches":
                self._send_json(_batch_list_payload())
            elif parsed.path == "/api/batch":
                self._send_json(
                    _batch_payload(
                        query.get("name", [""])[0],
                        query.get("q", [""])[0],
                        query.get("source", [""])[0],
                    )
                )
            elif parsed.path == "/api/document":
                self._send_json(
                    _document_payload(
                        query.get("batch", [""])[0],
                        query.get("id", [""])[0],
                    )
                )
            elif parsed.path == "/api/jobs":
                self._send_json(_jobs_payload())
            elif parsed.path == "/api/results":
                self._send_json(_results_payload())
            elif parsed.path == "/api/discovery/candidates":
                self._send_json(_discovery_candidates_payload())
            elif parsed.path.startswith("/api/discovery/candidates/"):
                source_id = parsed.path.rsplit("/", 1)[-1]
                self._send_json(_discovery_candidate_payload(source_id))
            elif parsed.path == "/api/discovery/portals":
                self._send_json(_discovery_portals_payload())
            elif parsed.path == "/api/discovery/keywords":
                self._send_json(_discovery_keywords_payload())
            elif parsed.path == "/api/discovery/reports":
                self._send_json(_discovery_reports_payload())
            elif parsed.path == "/api/workbook":
                self._send_json(
                    _workbook_payload(
                        query.get("name", [""])[0],
                        query.get("sheet", [""])[0],
                        query.get("q", [""])[0],
                    )
                )
            elif parsed.path == "/download/result":
                path = _named_file(
                    OUTPUT_DIRECTORY,
                    query.get("name", [""])[0],
                    "policy-opportunities-*",
                )
                self._send_file(path, attachment=True)
            elif parsed.path == "/download/job-log":
                job_id = query.get("id", [""])[0]
                path = _named_file(JOB_DIRECTORY, f"{job_id}.log", "*.log")
                self._send_file(path, attachment=True)
            elif parsed.path == "/download/raw":
                document = _document_payload(
                    query.get("batch", [""])[0],
                    query.get("id", [""])[0],
                )
                path = (PROJECT_ROOT / str(document["snapshot_path"])).resolve()
                if RAW_DIRECTORY.resolve() not in path.parents:
                    raise ValueError("原始快照路径不合法")
                self._send_file(path, attachment=True)
            else:
                self._send_json({"error": "页面不存在"}, HTTPStatus.NOT_FOUND)
        except FileNotFoundError as error:
            self._send_error_json(error, HTTPStatus.NOT_FOUND)
        except (OSError, TypeError, ValueError) as error:
            self._send_error_json(error, HTTPStatus.BAD_REQUEST)

    def do_PUT(self) -> None:
        try:
            path = urlparse(self.path).path
            if path not in {"/api/sources", "/api/prompts"}:
                self._send_json({"error": "接口不存在"}, HTTPStatus.NOT_FOUND)
                return
            payload = self._read_body()
            if path == "/api/sources":
                self._send_json(_save_sources(payload.get("sources")))
            else:
                self._send_json(_save_prompts(payload))
        except (json.JSONDecodeError, OSError, TypeError, ValueError) as error:
            self._send_error_json(error, HTTPStatus.BAD_REQUEST)

    def do_POST(self) -> None:
        try:
            payload = self._read_body()
            path = urlparse(self.path).path
            if path == "/api/collect":
                self._send_json(_start_collection(payload), HTTPStatus.ACCEPTED)
            elif path == "/api/analyze":
                self._send_json(_start_analysis(payload), HTTPStatus.ACCEPTED)
            elif path == "/api/stop-job":
                job_id = payload.get("job_id")
                if not job_id:
                    raise ValueError("job_id 不能为空")
                self._send_json(_stop_job(str(job_id)))
            elif path == "/api/discovery/search":
                self._send_json(_start_discovery_search(payload), HTTPStatus.ACCEPTED)
            elif path.startswith("/api/discovery/candidates/") and path.endswith("/promote"):
                source_id = path.split("/")[4]
                self._send_json(_discovery_promote(source_id, payload))
            elif path.startswith("/api/discovery/candidates/") and path.endswith("/review"):
                source_id = path.split("/")[4]
                self._send_json(_discovery_review(source_id, payload))
            else:
                self._send_json({"error": "接口不存在"}, HTTPStatus.NOT_FOUND)
        except (
            json.JSONDecodeError,
            OSError,
            TypeError,
            ValueError,
        ) as error:
            self._send_error_json(error, HTTPStatus.BAD_REQUEST)


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(prog="opportunity-radar-ui")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8501)
    parser.add_argument("--no-browser", action="store_true")
    return parser


def main() -> None:
    args = _parser().parse_args()
    if not 1 <= args.port <= 65535:
        raise SystemExit("port must be between 1 and 65535")
    server = ThreadingHTTPServer((args.host, args.port), RadarRequestHandler)
    url = f"http://{args.host}:{args.port}"
    print(f"AI 商机雷达控制台：{url}")
    if not args.no_browser and os.getenv("OPPORTUNITY_RADAR_UI_HEADLESS") != "1":
        threading.Timer(0.5, webbrowser.open, args=(url,)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n控制台已停止。")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
